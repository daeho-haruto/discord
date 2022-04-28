import discord
import asyncio
# from urllib import requsest
import urllib.parse
import urllib.request, re
from youtube_dl import YoutubeDL
import youtube_dl
# from selenium import webdriver
# from selenium.webdriver.chrome.options import Options
import openpyxl
import datetime
from datetime import datetime
import time
# from discord import FFmpegPCMAudio
from discord.utils import get
from discord.ext import commands
import random
# import bs4       
# import requests

token = "토큰" 
client = commands.Bot(command_prefix='!')
# intents = discord.Intents.default()
# client = discord.Client(intents=intents)

@client.event
async def on_ready():
    print("봇 준비 완료!")
    print(client.user.name)
    print("---------------")
    game = discord.Game("!명령어 라고 쳐주세요")
    await client.change_presence(status=discord.Status.online, activity=game)

@client.command()
async def 명령어(ctx):
    await ctx.send(embed = discord.Embed(title='명령어',description="""
!재생(url)  -> !음성채팅 입장 및 재생
!퇴장  -> !음성채팅 퇴장 (되도록 이걸로 퇴장좀..)
\n!기능   -> 기능추가 건의할 내용 작성
!기능보기  -> 기능목록 보기
!기능삭제  -> 기능삭제 (삭제할 내용 복붙)
!기능초기화  -> 기능목록 초기화
\n!알람추가  -> 원하는 월/일/시/분 알람 (미완성)
!알람보기  -> 알람목록 보기
!알람삭제   -> 알람삭제 (월/일/시/분 으로 삭제)
!알람초기화  -> 알람목록 초기화
\n!프로젝트  -> 동아리 프로젝트 
\n!분알람   -> 원하는 분 뒤에 알람
\n!타이머    -> 타이머 (초)
\n!투표  -> 공개투표용 (제목/투표1/투표2 ...)
\n!확인  -> 메세지를 확인했으면 이모티콘 클릭
\n!메세지삭제  -> 삭제 할 메세지 갯수
\n!관리자명령어 -> 관리자용 명령어 출력""", color = 0x00ff00))


@client.command()
async def 관리자명령어(ctx):
    i = (ctx.message.author.guild_permissions.administrator)
    if i is True:
        await ctx.send(embed = discord.Embed(title='관리자명령어',description="""
!경고       -> 한 대상에게 경고
!경고보기   -> 경고목록 출력 
!경고초기화 -> 경고목록 초기화""", color = 0x00ff00))
    if i is False:
        await ctx.message.channel.send("{}, 당신은 관리자 권한이 없습니다".format(ctx.message.author.mention))

@client.command()
async def 퇴장(ctx):
    global voice
    for vc in client.voice_clients:
        if vc.guild == ctx.message.guild:
            voice = vc
        
    await voice.disconnect()
    await ctx.message.channel.send("보이스채널 퇴장합니다.")

@client.command()
async def 재생(ctx):
    global voice
    await ctx.message.author.voice.channel.connect()
    await ctx.message.channel.send("보이스채널 입장합니다.")
    
    for vc in client.voice_clients:
        if vc.guild == ctx.message.guild:
            voice = vc

    url = ctx.message.content.split(" ")[1]
    await ctx.message.channel.send("잠시만 기다려주십시오.")
    option = {
        # 'outtmpl' : "file/" + url.split('=')[1] + ".mp3"
        'outtmpl' : "file/" + url.split('=')[1] + ".mp3",
        'format': 'bestaudio/best',
        'postprocessors': [{
            'key': 'FFmpegExtractAudio',
            'preferredcodec': 'mp3',
            'preferredquality': '192',
        }]
    }

    with youtube_dl.YoutubeDL(option) as ydl:
        ydl.download([url])
        info = ydl.extract_info(url, download=False)
        title = info["title"]

    voice.play(discord.FFmpegPCMAudio("file/" + url.split('=')[1] + ".mp3"))
    await ctx.message.channel.send(title + "을 재생합니다.")


@client.command()
async def 경고(ctx):
    i = (ctx.message.author.guild_permissions.administrator)
    if i is True:
        msg = ctx.message.content[4:]
        file = openpyxl.load_workbook("spam_경고.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(msg):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                file.save("spam_경고.xlsx")
                await ctx.message.channel.send("경고를 1회 받았습니다.")
                if sheet["B" + str(i)].value == 2:
                    await ctx.message.channel.send("누적 2회")
                elif sheet["B" + str(i)].value == 3:
                    await ctx.message.channel.send("누적 3회")
                elif sheet["B" + str(i)].value == 4:
                    await ctx.message.channel.send("누적 4회")
                elif sheet["B" + str(i)].value == 5:
                    await ctx.message.channel.send("누적 5회")
                elif sheet["B" + str(i)].value == 6:
                    await ctx.message.channel.send("누적 6회")
                elif sheet["B" + str(i)].value == 7:
                    await ctx.message.channel.send("누적 7회")
                elif sheet["B" + str(i)].value == 8:
                    await ctx.message.channel.send("누적 8회")
                elif sheet["B" + str(i)].value == 9:
                    await ctx.message.channel.send("누적 9회")
                elif sheet["B" + str(i)].value == 10:
                    await ctx.message.channel.send("누적 10회 퇴장조치")
                break
            if sheet["A" + str(i)].value == None:
                sheet["A" + str(i)].value = str(msg)
                sheet["B" + str(i)].value = 1
                file.save("spam_경고.xlsx")
                await ctx.message.channel.send("경고를 1회 받았습니다.")
                await ctx.message.channel.send("누적 1회")
                break
            i+=1
    if i is False:
        await ctx.message.channel.send("{}, 당신은 관리자 권한이 없습니다.".format(ctx.message.author.mention))

@client.command()
async def 경고보기(ctx):
    i = (ctx.message.author.guild_permissions.administrator)
    if i is True:
        file = openpyxl.load_workbook("spam_경고.xlsx")
        sheet = file.active
        i=1
        try:
            for i in range(1,100):
                await ctx.message.channel.send(sheet.cell(i,1).value)
                await ctx.message.channel.send(sheet.cell(i,2).value)
                i+=1
        except :
            await ctx.message.channel.send("--------")
            await ctx.message.channel.send("여기까지")
    if i is False:
        await ctx.message.channel.send("{}, 당신은 관리자 권한이 없습니다.".format(ctx.message.author.mention))


@client.command()
async def 경고초기화(ctx):
    i = (ctx.message.author.guild_permissions.administrator)
    if i is True:
        file = openpyxl.load_workbook("spam_경고.xlsx")
        sheet  = file.active
        sheet.delete_cols(1)
        sheet.delete_cols(1)
        file.save("spam_경고.xlsx")
        await ctx.message.channel.send("경고가 초기화 되었습니다.")
    if i is False:
        await ctx.message.channel.send("{}, 당신은 관리자 권한이 없습니다.".format(ctx.message.author.mention))



@client.command()
async def 기능(ctx):
    msg = ctx.message.content[4:]
    file = openpyxl.load_workbook("spam_기능.xlsx")
    sheet  = file.active
    i=1
    while True:
        if sheet["A" + str(i)].value == None:
            sheet["A" + str(i)].value = str(msg)
            file.save("spam_기능.xlsx")
            await ctx.message.channel.send(msg + " 기능이 추가 되었습니다.")
            break
        i+=1

@client.command()
async def 기능보기(ctx):
    file = openpyxl.load_workbook("spam_기능.xlsx")
    sheet = file.active
    i=1
    try:
        for i in range(1,100):
            await ctx.message.channel.send(sheet.cell(i,1).value)
            i+=1
    except :
        await ctx.message.channel.send("--------")
        await ctx.message.channel.send("여기까지")

@client.command()
async def 기능삭제(ctx):
    msg = ctx.message.content[6:]
    file = openpyxl.load_workbook("spam_기능.xlsx")
    sheet  = file.active
    i=1
    while True:
        if sheet["A" + str(i)].value == str(msg):
            sheet.delete_rows(i)
            file.save("spam_기능.xlsx")
            await ctx.message.channel.send(msg + " 기능이 삭제 되었습니다.")
            break
        i+=1

@client.command()
async def 기능초기화(ctx):
    file = openpyxl.load_workbook("spam_기능.xlsx")
    sheet  = file.active
    sheet.delete_cols(1)
    file.save("spam_기능.xlsx")
    await ctx.message.channel.send("기능이 초기화 되었습니다.")

@client.command()
async def 타이머(ctx):
    msg = ctx.message.content[5:]
    sec = int(msg)
    i=sec
    ms = await ctx.send(i)
    for i in range(sec, 0, -1):
        await asyncio.sleep(1) 
        await ms.edit(content=i-1)
    await ctx.message.channel.edit(content="종료")

@client.command()
async def 알람추가(ctx):
    msg = ctx.message.content[6:]
    file = openpyxl.load_workbook("spam_알람.xlsx")
    sheet  = file.active
    i=1
    while True:
        if sheet["A" + str(i)].value == None:
            li = msg.split("/")
            da1 = li.pop(0)
            da2 = li.pop(0)
            da3 = li.pop(0)
            da4 = li.pop(0)
            dd = da1 + da2 + da3 + da4
            sheet["A" + str(i)].value = str(dd)
            file.save("spam_알람.xlsx")
            month = dd[:2]
            day = dd[2:4]
            hour = dd[4:6]
            minu = dd[6:]
            await ctx.message.channel.send(month + '월 ' + day + "일 " + hour + '시 ' + minu + "분 알람이 추가되었습니다.")
            break
        i+=1
    while True:
        await asyncio.sleep(1800)
        global time
        datetime.today() 
        time = datetime.today().strftime("%m%d%H%M")  
        try:
            for i in range(1,100):
                dd = sheet.cell(i,1).value
                if dd == time :
                    month = dd[:2]
                    day = dd[2:4]
                    hour = dd[4:6]
                    minu = dd[6:]
                    sheet.delete_rows(i)
                    file.save("spam_알람.xlsx")
                    await ctx.message.channel.send(month + '월 ' + day + "일 " + hour + '시 ' + minu + "분 알람")
                else:
                    pass
                i+=1
        except:
            break
        

@client.command()
async def 알람삭제(ctx):
    msg = ctx.message.content[6:]
    file = openpyxl.load_workbook("spam_알람.xlsx")
    sheet  = file.active
    i=1
    while True:
        li = msg.split("/")
        da1 = li.pop(0)
        da2 = li.pop(0)
        da3 = li.pop(0)
        da4 = li.pop(0)
        dd = da1 + da2 + da3 + da4 
        if sheet["A" + str(i)].value == str(dd):
            sheet.delete_rows(i)
            file.save("spam_알람.xlsx")
            month = dd[:2]
            day = dd[2:4]
            hour = dd[4:6]
            minu = dd[6:]
            await ctx.message.channel.send(month + '월 ' + day + "일 " + hour + '시 ' + minu + "분 알람이 삭제 되었습니다.")
            break
        i+=1

@client.command()
async def 알람보기(ctx):
    file = openpyxl.load_workbook("spam_알람.xlsx")
    sheet  = file.active
    i=1
    try:
        for i in range(1,100):
            dd = sheet.cell(i,1).value
            month = dd[:2]
            day = dd[2:4]
            hour = dd[4:6]
            minu = dd[6:]
            await ctx.message.channel.send(month + '월 ' + day + "일 " + hour + '시 ' + minu + "분")
            i+=1
    except :
        await ctx.message.channel.send("--------")
        await ctx.message.channel.send("여기까지")

@client.command()
async def 알람초기화(ctx):
    file = openpyxl.load_workbook("spam_알람.xlsx")
    sheet  = file.active
    sheet.delete_cols(1)
    file.save("spam_알람.xlsx")
    await ctx.message.channel.send("알람이 초기화 되었습니다.")
    
@client.command()
async def 메세지삭제(ctx):
    number = int(ctx.message.content.split()[1])
    await ctx.message.channel.purge(limit=number + 1)
    await ctx.message.channel.send(f"{number}개 메세지 삭제완료")
    

@client.command()
async def 투표(ctx):
    vote = ctx.message.content[4:].split("/")
    await ctx.message.channel.send("★투표 - " + vote[0])
    for i in range(1, len(vote)):
        choose = await ctx.message.channel.send("```" + vote[i] + "```")
        await choose.add_reaction('👍')

@client.command()
async def 분알람(ctx):
    msg = ctx.message.content[5:]
    await ctx.message.channel.send(msg + "분 후에 울립니다.")
    msg = int(msg)
    await asyncio.sleep(60*msg) 
    msg = str(msg)
    await ctx.message.channel.purge(limit=2)
    await ctx.message.channel.send(msg + "분 알람")

@client.command()
async def 가위바위보(ctx):
	if ctx.message.content.startswith('!가위바위보'):
		rsp = ["가위","바위","보"]
		embed = discord.Embed(title="가위바위보",description="가위바위보를 합니다 5초내로 (가위/바위/보)를 써주세요!", color=0x00aaaa)
		channel = ctx.message.channel
		msg1 = await ctx.message.channel.send(embed=embed)
		def check(m):
			return m.author == ctx.message.author and m.channel == channel
		try:
			msg2 = await client.wait_for('message', timeout=5.0, check=check)
		except asyncio.TimeoutError:
			await msg1.delete()
			embed = discord.Embed(title="가위바위보",description="앗 5초가 지났네요...!", color=0x00aaaa)
			await ctx.message.channel.send(embed=embed)
			return
		else:
			await msg1.delete()
			bot_rsp = str(random.choice(rsp))
			user_rsp  = str(msg2.content)
			answer = ""
			if bot_rsp == user_rsp:
				answer = "저는 " + bot_rsp + "을 냈고, 당신은 " + user_rsp + "을 내셨내요.\n" + "아쉽지만 비겼습니다."
			elif (bot_rsp == "가위" and user_rsp == "바위") or (bot_rsp == "보" and user_rsp == "가위") or (bot_rsp == "바위" and user_rsp == "보"):
				answer = "저는 " + bot_rsp + "을 냈고, 당신은 " + user_rsp + "을 내셨내요.\n" + "아쉽지만 제가 졌습니다."
			elif (bot_rsp == "바위" and user_rsp == "가위") or (bot_rsp == "가위" and user_rsp == "보") or (bot_rsp == "보" and user_rsp == "바위"):
				answer = "저는 " + bot_rsp + "을 냈고, 당신은 " + user_rsp + "을 내셨내요.\n" + "제가 이겼습니다!"
			else:
				embed = discord.Embed(title="가위바위보",description="앗, 가위, 바위, 보 중에서만 내셔야죠...", color=0x00aaaa)
				await ctx.message.channel.send(embed=embed)
				return
			embed = discord.Embed(title="가위바위보",description=answer, color=0x00aaaa)
			await ctx.message.channel.send(embed=embed)
			return

@client.command()
async def 확인(ctx):
    choose = await ctx.message.channel.send("확인하면 눌러주세요")
    await choose.add_reaction('👍')

@client.command()
async def 공지(ctx):
	if ctx.message.content.startswith('!공지'):
		
		embed = discord.Embed(title="공지",description="5분내로 공지를 작성해주세요", color=0x00aaaa)
		channel = ctx.message.channel
		msg1 = await ctx.message.channel.send(embed=embed)
		def check(m):
			return m.author == ctx.message.author and m.channel == channel
		try:
			msg2 = await client.wait_for('message', timeout=300.0, check=check)
		except asyncio.TimeoutError:
			await msg1.delete()
			embed = discord.Embed(title="공지",description="앗 5분이 지났네요...!", color=0x00aaaa)
			await ctx.message.channel.send(embed=embed)
			return
		else:
			await msg1.delete()
			await ctx.message.channel.purge(limit= 2)
			user_rsp  = str(msg2.content)
			answer = ""
			embed = discord.Embed(title="공지",description=user_rsp+"@here", color=0x00aaaa)
			await ctx.message.channel.send(embed=embed)
			return

@client.command()
async def 프로젝트(ctx):
    global emo
    embed = discord.Embed(title = "프로젝트", color = 0x00ff00)
    embed.add_field(name = "OPENCV_RULER 🔴", value = "OPENCV 를 이용하여 물체 길이 측정", inline = False)
    embed.add_field(name = "PYTHON_GAME 🟠", value = "PYGAME 을 이용하여 각자 게임 제작", inline = False)
    embed.add_field(name = "PYTHON_DISCORDBOT 🟡", value = "PYTHON 을 이용한 디스코드봇 제작", inline = False)
    embed.add_field(name = "UNITY 🟢", value = "UNITY 를 이용하여 게임 제작", inline = False)
    embed.add_field(name = "SPAM_WEBSEVER 🔵", value = "SPAM 동아리 웹페이지 제작", inline = False)
    emo = await ctx.send(embed = embed)
    await emo.add_reaction("🔴")
    await emo.add_reaction("🟠")
    await emo.add_reaction("🟡")
    await emo.add_reaction("🟢")
    await emo.add_reaction("🔵")

@client.event
async def on_reaction_add(reaction, users):
    if users.bot == 1:
        pass
    else:
        try:
            await emo.delete()
        except :
            pass
        else :
            if str(reaction.emoji) == '🔴':
                await reaction.message.channel.send(embed = discord.Embed(title='OPENCV_RULER',description="", color = 0x00ff00))
                await reaction.message.channel.send("https://drive.google.com/file/d/1W6_nhopj-pOT52I5J7mMVN2fhvrcf5RI/view?usp=sharing")
            elif str(reaction.emoji) == '🟠':
                await reaction.message.channel.send(embed = discord.Embed(title='PYTHON_GAME',description="", color = 0x00ff00))
                await reaction.message.channel.send("https://drive.google.com/file/d/1l8KPKG6RORcx0Kf0eL8_e3uvdd0noyOB/view?usp=sharing")
            elif str(reaction.emoji) == '🟡':
                await reaction.message.channel.send(embed = discord.Embed(title='PYTHON_DISCORDBOT',description="", color = 0x00ff00))
                # await reaction.message.channel.send("https://drive.google.com/drive/u/2/folders/1q9aBThAK_K6TvtFGtHfSh4huUlK-NMYB")
            elif str(reaction.emoji) == '🟢':
                await reaction.message.channel.send(embed = discord.Embed(title='UNITY',description="", color = 0x00ff00))
                await reaction.message.channel.send("https://drive.google.com/file/d/1iVWDq2HcairKLNSyIs8Xp32Epag4MvyA/view?usp=sharing")
            elif str(reaction.emoji) == '🔵':
                await reaction.message.channel.send(embed = discord.Embed(title='SPAM_WEBSEVER',description="", color = 0x00ff00))
                # await reaction.message.channel.send("https://drive.google.com/drive/u/2/folders/13bF8MM1pGANKz2GNKeoP7rA6HCMmMETZ")

client.run(token)
